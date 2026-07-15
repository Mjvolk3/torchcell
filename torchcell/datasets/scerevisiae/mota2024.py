# torchcell/datasets/scerevisiae/mota2024
# [[torchcell.datasets.scerevisiae.mota2024]]
# https://github.com/Mjvolk3/torchcell/tree/main/torchcell/datasets/scerevisiae/mota2024
# Test file: tests/torchcell/datasets/scerevisiae/test_mota2024.py
"""Mota 2024 acetic/butyric/octanoic acid chemogenomic screen (env x geno -> response).

Mota et al. 2024 (Microbial Cell Factories, doi:10.1186/s12934-024-02309-0; PMC10903034)
screened the ~5100-strain S. cerevisiae BY4741 Euroscarf HAPLOID single-deletion collection
for susceptibility to three linear monocarboxylic weak acids at EQUIVALENT moderate
inhibitory concentrations -- 75 mM acetic acid (C2), 14 mM butyric acid (C4), and 0.30 mM
octanoic acid (C8) -- on YPD solid medium at pH 4.5, 30 C, scored by spot assay after 48 h
against the parental strain. This is a plain haploid deletion collection (NOT a HIP/HOP
heterozygous screen), so every screened strain is a single-gene KanMX deletion in BY4741 and
there is no engineered-heterozygous / dosage complication.

The readout is a CATEGORICAL susceptibility call relative to the parental strain (paper
Fig. S2 / Table 1 caption): ``+`` = minor-to-moderate growth inhibition, ``++`` = total
growth inhibition (no growth), ``0`` = no detectable susceptibility phenotype. The paper only
enumerates the SUSCEPTIBLE mutants (score + or ++) in Additional files 1-3 (Tables S1/S2/S3);
non-susceptible strains (score 0) are NOT listed, so this dataset carries one categorical
record per (susceptible deletion x acid) pair. No mutant conferred INCREASED tolerance (the
paper reports none), so there is no "resistant" category.

This maps onto the WS15 environment-perturbation ontology: ``EnvironmentResponseExperiment``
= single-deletion ``Genotype`` (a ``KanMxDeletionPerturbation`` in the BY4741 reference
background) x aerobic ``Environment`` carrying a ``SmallMoleculePerturbation`` (the acid, a
typed ``Compound``, at its fixed inhibitory molar concentration) ->
``EnvironmentResponsePhenotype`` (``measurement_type=categorical``, ``category`` the
susceptibility call). The parental BY4741 baseline (no susceptibility) is the reference.

PROVENANCE / SOURCING (all sha256-pinned; see class constants + provenance record):
- Concentrations 75/14/0.30 mM at pH 4.5, 30 C, 48 h spot assay, BY4741 Euroscarf haploid
  collection: Methods "Genome-wide search..." + "Selection of equivalent concentrations...".
- Categorical scoring (+/++/0 definitions): Methods + Fig. S2 + Table S1/S2/S3 captions.
- The susceptible-gene lists WITH per-gene scores come from the BMC open-access
  supplementary spreadsheets (Additional file 1 = acetic Table S1, Additional file 2 =
  butyric Table S2, Additional file 3 = octanoic Table S3), downloaded from
  static-content.springer.com (scriptable) and sha256-pinned. Parsed record counts match the
  paper's reported totals exactly (acetic 377 = 46 ++ + 331 +; butyric 422 = 51 ++ + 371 +;
  octanoic 490 = 53 ++ + 437 +).
- N_SAMPLES / replicate structure for the genome-wide spot screen is NOT reported by the
  paper (the ">= 3 independent experiments" statements apply only to the CFU-viability and
  intracellular-pH physiological assays, NOT the disruptome screen). Per the source-or-stop
  policy it is therefore left ``None`` (categorical readout carries no SE anyway) and flagged.

SOURCE ARTIFACTS / QUIRKS handled deterministically:
- RNR4 (YGR180C) is listed TWICE in every table (a source duplicate the paper's headline
  totals count); for acetic/butyric both rows are ``+``, for octanoic the two rows conflict
  (``+`` and ``++``). Records are deduplicated per (resolved ORF, acid), keeping the MORE
  SEVERE score (``++`` > ``+``), so RNR4 is one record per acid (octanoic -> ``++``).
- Six gene tokens do not resolve to an SGD R64 systematic ORF via the genome alias table and
  are DROPPED (never guessed): REF1, RLM2, SBR2 (all three acids), ILM2 (butyric), VPS236
  (butyric, octanoic), SIW15 (octanoic) -- 13 records total. Final: acetic 373, butyric 416,
  octanoic 484 = 1273 records. Dropped tokens are logged; resolving them is a flagged
  follow-up (likely SI typos: VPS236->VPS36?, SIW15->SIW14?, ILM2->ILM1?, RLM2->RLM1?).
"""

import hashlib
import logging
import os
import os.path as osp
import pickle
import re
import urllib.request
from collections.abc import Callable
from typing import Any

import lmdb
import openpyxl
from tqdm import tqdm

from torchcell.data import ExperimentDataset, post_process
from torchcell.datamodels.schema import (
    Compound,
    Concentration,
    ConcentrationUnit,
    Environment,
    EnvironmentResponseExperiment,
    EnvironmentResponseExperimentReference,
    EnvironmentResponsePhenotype,
    Experiment,
    ExperimentReference,
    Genotype,
    KanMxDeletionPerturbation,
    MeasurementType,
    Media,
    Publication,
    ReferenceGenome,
    SmallMoleculePerturbation,
    Temperature,
)
from torchcell.datasets.dataset_registry import register_dataset
from torchcell.sequence.genome.scerevisiae import SCerevisiaeGenome

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

DOI = "10.1186/s12934-024-02309-0"
# BMC open-access supplementary spreadsheets (scriptable static-content mirror).
_ESM_URL = (
    "https://static-content.springer.com/esm/"
    "art%3A10.1186%2Fs12934-024-02309-0/MediaObjects/"
    "12934_2024_2309_MOESM{n}_ESM.xlsx"
)

# Susceptibility symbols -> (severity rank, category string). Severity resolves the
# RNR4 source duplicate (keep the more severe score). Category strings are verbatim
# semantics from the paper's Fig. S2 / Table caption definitions.
_SEVERITY = {"+": 1, "++": 2}
_CATEGORY = {
    "+": "minor_to_moderate_growth_inhibition",
    "++": "total_growth_inhibition",
}
_REFERENCE_CATEGORY = "no_detectable_susceptibility"

MEASUREMENT_UNITS = (
    "spot-assay susceptibility call vs parental BY4741 after 48 h on YPD (pH 4.5): "
    "'minor_to_moderate_growth_inhibition' (+) or 'total_growth_inhibition' (++)"
)

_SYSTEMATIC_RE = re.compile(r"^Y[A-P][LR]\d{3}[WC](-[A-Z])?$")

# Per-acid source spec: MOESM index, filename, sha256, compound, molar concentration (mM).
_ACID_SPECS: list[dict[str, Any]] = [
    {
        "acid": "acetic",
        "n": 1,
        "filename": "12934_2024_2309_MOESM1_ESM.xlsx",
        "sha256": "b23ad28141e70b307048fc69475aedd4e3cf880118ae9d0d806b6d9f91205e42",
        "compound_name": "acetic acid",
        "concentration_mM": 75.0,
    },
    {
        "acid": "butyric",
        "n": 2,
        "filename": "12934_2024_2309_MOESM2_ESM.xlsx",
        "sha256": "a7a1aaee1c76e52d8fe435326790c89170ab43ec96b92ef272903d8e78a1e81f",
        "compound_name": "butyric acid",
        "concentration_mM": 14.0,
    },
    {
        "acid": "octanoic",
        "n": 3,
        "filename": "12934_2024_2309_MOESM3_ESM.xlsx",
        "sha256": "27f1508641ad5e7cc29ab8611739d4940da355c1dffbe9c9a908c267cdf5d455",
        "compound_name": "octanoic acid",
        "concentration_mM": 0.30,
    },
]


@register_dataset
class EnvChemgenMota2024Dataset(ExperimentDataset):
    """Acetic/butyric/octanoic acid chemogenomic env x geno -> categorical susceptibility."""

    def __init__(
        self,
        root: str = "data/torchcell/env_chemgen_mota2024",
        io_workers: int = 0,
        genome: SCerevisiaeGenome | None = None,
        transform: Callable[..., Any] | None = None,
        pre_transform: Callable[..., Any] | None = None,
        **kwargs: Any,
    ) -> None:
        """Initialize the dataset; a genome is REQUIRED for common-name -> ORF mapping."""
        self.genome = genome
        super().__init__(root, io_workers, transform, pre_transform, **kwargs)

    @property
    def experiment_class(self) -> type[Experiment]:
        """Experiment schema class produced by this dataset."""
        return EnvironmentResponseExperiment

    @property
    def reference_class(self) -> type[ExperimentReference]:
        """Experiment-reference schema class produced by this dataset."""
        return EnvironmentResponseExperimentReference

    @property
    def raw_file_names(self) -> list[str]:
        """The three BMC supplementary spreadsheets required before processing."""
        return [spec["filename"] for spec in _ACID_SPECS]

    def download(self) -> None:
        """Download the three BMC supplementary xlsx files; verify each pinned sha256."""
        os.makedirs(self.raw_dir, exist_ok=True)
        for spec in _ACID_SPECS:
            dest = osp.join(self.raw_dir, spec["filename"])
            if osp.exists(dest):
                continue
            url = _ESM_URL.format(n=spec["n"])
            log.info("Downloading Mota2024 supplement %s from %s", spec["acid"], url)
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=180) as resp:
                data = resp.read()
            digest = hashlib.sha256(data).hexdigest()
            if digest != spec["sha256"]:
                raise RuntimeError(
                    f"{spec['filename']} sha256 mismatch: got {digest}, "
                    f"expected {spec['sha256']}"
                )
            with open(dest, "wb") as handle:
                handle.write(data)
            log.info("Wrote %s (%d bytes, sha256 verified)", dest, len(data))

    def _resolver(self) -> Callable[[str], str | None]:
        """Build a common-name/ORF -> systematic-ORF resolver from the injected genome."""
        if self.genome is None:
            raise RuntimeError(
                "EnvChemgenMota2024Dataset requires a genome for gene-name resolution; "
                "inject SCerevisiaeGenome(...)"
            )
        genome = self.genome
        df = genome.gene_attribute_table
        alias_map = genome.alias_to_systematic

        def resolve(token: str) -> str | None:
            gene = token.upper()
            if _SYSTEMATIC_RE.match(gene):
                return gene
            hit = df[df["gene"] == gene]
            if not hit.empty:
                return str(hit.iloc[0]["ID"])
            hit = df[df["Alias"] == gene]
            if not hit.empty:
                return str(hit.iloc[0]["ID"])
            candidates = alias_map.get(gene, [])
            if candidates:
                return candidates[0]
            return None

        return resolve

    def _parse_acid(
        self, spec: dict[str, Any], resolve: Callable[[str], str | None]
    ) -> dict[str, tuple[str, str]]:
        """Parse one acid spreadsheet into {ORF: (score_symbol, source_token)}.

        Deduplicates per resolved ORF keeping the more severe score (resolves the RNR4
        source duplicate). Unresolvable tokens are skipped and logged.
        """
        path = osp.join(self.raw_dir, spec["filename"])
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            i for i, r in enumerate(rows) if r and r[0] == "Gene/ORF name"
        )
        best: dict[str, tuple[str, str]] = {}
        dropped: list[str] = []
        raw = 0
        for row in rows[header_idx + 1 :]:
            if not row or row[0] is None or row[2] is None:
                continue
            token = str(row[0]).replace("\xa0", " ").strip()
            score = str(row[2]).strip()
            if not token or score not in _SEVERITY:
                continue
            raw += 1
            orf = resolve(token)
            if orf is None:
                dropped.append(token)
                continue
            if orf not in best or _SEVERITY[score] > _SEVERITY[best[orf][0]]:
                best[orf] = (score, token)
        log.info(
            "Mota2024 %s acid: %d raw susceptible rows -> %d records "
            "(dropped %d unresolvable: %s)",
            spec["acid"],
            raw,
            len(best),
            len(set(dropped)),
            sorted(set(dropped)),
        )
        return best

    def _environment(self, spec: dict[str, Any]) -> Environment:
        """Aerobic YPD (pH 4.5) plate carrying the acid at its fixed inhibitory molarity."""
        return Environment(
            media=Media(name="YPD, pH 4.5", state="solid", is_synthetic=False),
            temperature=Temperature(value=30.0),
            perturbations=[
                SmallMoleculePerturbation(
                    compound=Compound(name=spec["compound_name"]),
                    concentration=Concentration(
                        value=spec["concentration_mM"],
                        unit=ConcentrationUnit.millimolar,
                    ),
                )
            ],
            aerobicity="aerobic",
            duration_hours=48.0,
        )

    def _reference(
        self, environment: Environment
    ) -> EnvironmentResponseExperimentReference:
        """Parental BY4741 baseline: no detectable susceptibility in the same acid medium."""
        phenotype_reference = EnvironmentResponsePhenotype(
            measurement_type=MeasurementType.categorical,
            category=_REFERENCE_CATEGORY,
            units=MEASUREMENT_UNITS,
        )
        return EnvironmentResponseExperimentReference(
            dataset_name=self.name,
            genome_reference=ReferenceGenome(
                species="Saccharomyces cerevisiae", strain="BY4741"
            ),
            environment_reference=environment.model_copy(),
            phenotype_reference=phenotype_reference,
        )

    def _experiment(
        self, *, orf: str, token: str, score: str, environment: Environment
    ) -> EnvironmentResponseExperiment:
        """Build one env x geno -> categorical-susceptibility experiment for (gene, acid)."""
        genotype = Genotype(
            perturbations=[
                KanMxDeletionPerturbation(
                    systematic_gene_name=orf, perturbed_gene_name=token
                )
            ]
        )
        phenotype = EnvironmentResponsePhenotype(
            measurement_type=MeasurementType.categorical,
            category=_CATEGORY[score],
            units=MEASUREMENT_UNITS,
        )
        return EnvironmentResponseExperiment(
            dataset_name=self.name,
            genotype=genotype,
            environment=environment,
            phenotype=phenotype,
        )

    @post_process
    def process(self) -> None:
        """Parse the three acid spreadsheets into categorical records; write LMDB."""
        resolve = self._resolver()
        publication = Publication(doi=DOI, doi_url=f"https://doi.org/{DOI}")
        pub_dump = publication.model_dump()

        os.makedirs(self.preprocess_dir, exist_ok=True)
        os.makedirs(self.processed_dir, exist_ok=True)
        env = lmdb.open(osp.join(self.processed_dir, "lmdb"), map_size=int(1e11))
        idx = 0
        with env.begin(write=True) as txn:
            for spec in _ACID_SPECS:
                environment = self._environment(spec)
                reference = self._reference(environment)
                ref_dump = reference.model_dump()
                orf_scores = self._parse_acid(spec, resolve)
                for orf, (score, token) in tqdm(
                    sorted(orf_scores.items()), desc=f"{spec['acid']} acid"
                ):
                    experiment = self._experiment(
                        orf=orf, token=token, score=score, environment=environment
                    )
                    txn.put(
                        f"{idx}".encode(),
                        pickle.dumps(
                            {
                                "experiment": experiment.model_dump(),
                                "reference": ref_dump,
                                "publication": pub_dump,
                            }
                        ),
                    )
                    idx += 1
        env.close()
        log.info("Wrote %d Mota2024 environment-response experiments to LMDB", idx)

    def preprocess_raw(self, df: Any, preprocess: dict[str, Any] | None = None) -> Any:
        """Preprocessing is handled inside process() for this dataset."""
        return df

    def create_experiment(self) -> None:
        """Experiment construction is handled inline in process() for this dataset."""
        raise NotImplementedError


def main() -> None:
    """Build/load the dataset for interactive debugging."""
    from dotenv import load_dotenv

    load_dotenv()
    data_root = os.environ["DATA_ROOT"]
    genome = SCerevisiaeGenome(
        genome_root=osp.join(data_root, "data/sgd/genome"),
        go_root=osp.join(data_root, "data/go"),
        overwrite=False,
    )
    root = osp.join(data_root, "data/torchcell/env_chemgen_mota2024")
    dataset = EnvChemgenMota2024Dataset(root=root, genome=genome)
    print(f"len = {len(dataset)}")
    print(dataset[0])


if __name__ == "__main__":
    main()
